from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.http import JsonResponse

from django.shortcuts import render
from django.utils.dateparse import parse_date
from .models import MouvementCaisse, Parametre

from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa

from django.shortcuts import redirect
from .forms import MouvementCaisseForm
from django.contrib import messages

from django.core.paginator import Paginator
from django.db.models import Sum, Q
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from openpyxl.utils import get_column_letter

from .models import MouvementCaisse, Caisse, TypeMouvement, ModePaiement, Parametre

from django.contrib.auth.decorators import login_required, user_passes_test

def export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relevé de caisse"

    # --- Styles ---
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # --- En-têtes ---
    headers = [
        "Date", "N° pièce", "Libellé", "Type", "Caisse",
        "Mode paiement", "Entrée", "Sortie", "Solde cumulé"
    ]

    ws.append(headers)

    # Appliquer styles sur les headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = bold
        cell.alignment = center
        cell.border = thin_border

    # --- Données ---
    mouvements = MouvementCaisse.objects.order_by("date_mouvement", "id")

    solde = 0
    row_num = 2

    for mv in mouvements:
        entree = mv.entree or 0
        sortie = mv.sortie or 0
        solde += entree - sortie

        row = [
            mv.date_mouvement.strftime("%d/%m/%Y"),
            mv.num_piece,
            mv.libelle,
            mv.type_mouvement.libelle,
            mv.caisse.libelle,
            mv.mode_paiement.libelle,
            entree,
            sortie,
            solde,
        ]

        ws.append(row)

        # Styles ligne par ligne
        for col in range(1, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            if col >= 7:  # Montants à droite
                cell.alignment = right

        row_num += 1

    # --- Ajustement automatique des colonnes ---
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # --- Préparer la réponse HTTP ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="releve_caisse.xlsx"'

    wb.save(response)
    return response

def is_caissier(user):
    return user.is_staff or user.groups.filter(name="Caissier").exists()
@login_required
def dashboard(request):
    ...

@login_required
def suivi_caisse(request):
    ...

@login_required
def recap_mensuel(request):
    ...

@login_required
@user_passes_test(is_caissier)
def mouvement_create(request):
    ...


def api_mensuel(request):
    """
    Retourne les totaux mensuels d'entrées/sorties au format JSON
    pour alimenter Chart.js.
    """
    annee = request.GET.get("annee")
    qs = MouvementCaisse.objects.all()
    if annee:
        qs = qs.filter(date_mouvement__year=int(annee))

    agg = (
        qs.annotate(mois=TruncMonth("date_mouvement"))
        .values("mois")
        .annotate(entree=Sum("entree"), sortie=Sum("sortie"))
        .order_by("mois")
    )

    data = []
    for row in agg:
        data.append({
            "mois": row["mois"].strftime("%Y-%m") if row["mois"] else "",
            "entree": float(row["entree"] or 0),
            "sortie": float(row["sortie"] or 0),
        })

    return JsonResponse({"data": data})

def mouvement_create(request):
    if request.method == "POST":
        form = MouvementCaisseForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Mouvement enregistré avec succès.")
            return redirect("suivi_caisse")
        else:
            messages.error(request, "Veuillez corriger les erreurs.")
    else:
        form = MouvementCaisseForm()

    return render(request, "caisse/mouvement_form.html", {"form": form})



def suivi_caisse(request):
    """
    Liste des mouvements de caisse avec filtres + totaux + pagination.
    """

    # --- Filtres GET ---
    date_debut = request.GET.get("date_debut") or ""
    date_fin = request.GET.get("date_fin") or ""
    caisse_id = request.GET.get("caisse") or ""
    type_id = request.GET.get("type") or ""
    mode_id = request.GET.get("mode") or ""
    q = request.GET.get("q") or ""

    mouvements = (
        MouvementCaisse.objects
        .select_related("caisse", "type_mouvement", "mode_paiement")
        .order_by("-date_mouvement", "-id")
    )

    # Filtre période
    if date_debut:
        mouvements = mouvements.filter(date_mouvement__gte=date_debut)
    if date_fin:
        mouvements = mouvements.filter(date_mouvement__lte=date_fin)

    # Filtre caisse
    if caisse_id:
        mouvements = mouvements.filter(caisse_id=caisse_id)

    # Filtre type de mouvement
    if type_id:
        mouvements = mouvements.filter(type_mouvement_id=type_id)

    # Filtre mode de paiement
    if mode_id:
        mouvements = mouvements.filter(mode_paiement_id=mode_id)

    # Recherche texte
    if q:
        mouvements = mouvements.filter(
            Q(libelle__icontains=q)
            | Q(num_piece__icontains=q)
            | Q(observations__icontains=q)
        )

    # --- Totaux sur le filtre courant ---
    agg = mouvements.aggregate(
        total_entree=Sum("entree"),
        total_sortie=Sum("sortie"),
    )
    total_entree = agg["total_entree"] or 0
    total_sortie = agg["total_sortie"] or 0
    solde_filtre = total_entree - total_sortie

    # Paramètres (solde initial)
    param = Parametre.objects.first()
    solde_initial = param.solde_initial if param else 0
    solde_global = solde_initial + solde_filtre

    # --- Pagination ---
    paginator = Paginator(mouvements, 25)  # 25 lignes par page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "mouvements": page_obj,   # alias pratique
        "paginator": paginator,

        # Filtres courants (pour garder les valeurs dans le formulaire)
        "date_debut": date_debut,
        "date_fin": date_fin,
        "caisse_id": caisse_id,
        "type_id": type_id,
        "mode_id": mode_id,
        "q": q,

        # Listes pour les <select>
        "caisses": Caisse.objects.all(),
        "types": TypeMouvement.objects.all(),
        "modes": ModePaiement.objects.all(),

        # Totaux
        "total_entree": total_entree,
        "total_sortie": total_sortie,
        "solde_filtre": solde_filtre,
        "solde_initial": solde_initial,
        "solde_global": solde_global,
    }

    return render(request, "caisse/suivi_caisse.html", context)


def recap_mensuel(request):
    # Année sélectionnée (par défaut : dernière année des mouvements ou des paramètres)
    annee_str = request.GET.get("annee")

    param = Parametre.objects.order_by("-annee_analyse").first()
    annee_defaut = param.annee_analyse if param else None

    if annee_str:
        annee = int(annee_str)
    else:
        annee = annee_defaut

    mouvements = MouvementCaisse.objects.all()
    if annee:
        mouvements = mouvements.filter(date_mouvement__year=annee)

    # Regroupement par mois
    agr = (
        mouvements
        .annotate(mois=TruncMonth("date_mouvement"))
        .values("mois")
        .annotate(
            total_entree=Sum("entree"),
            total_sortie=Sum("sortie"),
        )
        .order_by("mois")
    )

    # Calcul du solde cumulé par mois
    solde_initial = param.solde_initial if param else 0
    solde_courant = solde_initial

    lignes = []
    for row in agr:
        entree = row["total_entree"] or 0
        sortie = row["total_sortie"] or 0
        solde_courant += entree - sortie
        lignes.append(
            {
                "mois": row["mois"],
                "total_entree": entree,
                "total_sortie": sortie,
                "solde_cumule": solde_courant,
            }
        )

    context = {
        "annee": annee,
        "annee_defaut": annee_defaut,
        "lignes": lignes,
        "parametre": param,
        "solde_initial": solde_initial,
    }
    return render(request, "caisse/recap_mensuel.html", context)

from django.db.models.functions import TruncMonth


from django.shortcuts import render
from django.db.models import Sum
from .models import MouvementCaisse

def dashboard(request):
    # Base queryset
    qs = MouvementCaisse.objects.select_related("caisse", "type_mouvement", "mode_paiement")

    # --- KPI globaux ---
    total_entree = qs.aggregate(total=Sum("entree"))["total"] or 0
    total_sortie = qs.aggregate(total=Sum("sortie"))["total"] or 0

    # solde initial (si plusieurs Parametre, on prend le premier)
    param = Parametre.objects.first()
    solde_initial = param.solde_initial if param else 0
    solde_final = solde_initial + total_entree - total_sortie

    nb_mouvements = qs.count()

    # Derniers mouvements pour le tableau en bas
    derniers = qs.order_by("-date_mouvement", "-id")[:5]

    # --- Séries mensuelles pour graphique Entrée / Sortie ---
    monthly_qs = (
        qs.annotate(month=TruncMonth("date_mouvement"))
          .values("month")
          .annotate(
              entree=Sum("entree"),
              sortie=Sum("sortie")
          )
          .order_by("month")
    )

    # On transforme en listes Python (dernier 6 mois par exemple)
    monthly_data = list(monthly_qs)[-6:]

    chart_labels = [
        m["month"].strftime("%m/%Y") if m["month"] else "N/A"
        for m in monthly_data
    ]
    chart_entrees = [float(m["entree"] or 0) for m in monthly_data]
    chart_sorties = [float(m["sortie"] or 0) for m in monthly_data]

    # --- Répartition par caisse (graphique donut) ---
    by_caisse_qs = (
        qs.values("caisse__libelle")
          .annotate(solde=Sum("entree") - Sum("sortie"))
          .order_by("-solde")
    )

    caisse_labels = [row["caisse__libelle"] or "Sans caisse" for row in by_caisse_qs]
    caisse_values = [float(row["solde"] or 0) for row in by_caisse_qs]

    # --- KPI du mois courant ---
    today = timezone.now().date()
    first_day = today.replace(day=1)
    courant = qs.filter(date_mouvement__gte=first_day, date_mouvement__lte=today)

    mois_entree = courant.aggregate(total=Sum("entree"))["total"] or 0
    mois_sortie = courant.aggregate(total=Sum("sortie"))["total"] or 0
    mois_solde = mois_entree - mois_sortie

    context = {
        # KPI globaux
        "solde_initial": solde_initial,
        "total_entree": total_entree,
        "total_sortie": total_sortie,
        "solde_final": solde_final,
        "nb_mouvements": nb_mouvements,

        # KPI du mois
        "mois_entree": mois_entree,
        "mois_sortie": mois_sortie,
        "mois_solde": mois_solde,
        "mois_courant": today.strftime("%B %Y"),

        # Derniers mouvements
        "derniers": derniers,

        # Données pour graphiques
        "chart_labels": chart_labels,
        "chart_entrees": chart_entrees,
        "chart_sorties": chart_sorties,
        "caisse_labels": caisse_labels,
        "caisse_values": caisse_values,
    }

    return render(request, "caisse/dashboard.html", context)


def api_mensuel(request):
    """
    Renvoie un JSON du type :
    {
      "data": [
        {"mois": "01/2025", "entree": 12345, "sortie": 6789},
        ...
      ]
    }
    """
    qs = (
        MouvementCaisse.objects
        .annotate(mois=TruncMonth("date_mouvement"))
        .values("mois")
        .annotate(
            total_entree=Sum("entree"),
            total_sortie=Sum("sortie"),
        )
        .order_by("mois")
    )

    data = []
    for row in qs:
        mois = row["mois"]
        if mois is None:
            continue
        data.append(
            {
                "mois": mois.strftime("%m/%Y"),
                "entree": float(row["total_entree"] or 0),
                "sortie": float(row["total_sortie"] or 0),
            }
        )

    return JsonResponse({"data": data})

def mouvement_create(request):
    if request.method == "POST":
        form = MouvementCaisseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("suivi_caisse")
    else:
        form = MouvementCaisseForm()

    context = {
        "form": form,
    }
    return render(request, "caisse/mouvement_form.html", context)

    # Derniers mouvements
    derniers = mvt.select_related("caisse", "type_mouvement").order_by("-date_mouvement")[:5]

    # Récap mois courant
    from datetime import date
    today = date.today()
    mois_courant = mvt.filter(
        date_mouvement__year=today.year,
        date_mouvement__month=today.month,
    )
    mois_entree = mois_courant.aggregate(Sum("entree"))["entree__sum"] or 0
    mois_sortie = mois_courant.aggregate(Sum("sortie"))["sortie__sum"] or 0

    context = {
        "total_entree": total_entree,
        "total_sortie": total_sortie,
        "solde": solde,
        "derniers": derniers,
        "mois_entree": mois_entree,
        "mois_sortie": mois_sortie,
        "mois_solde": mois_entree - mois_sortie,
    }
    return render(request, "caisse/dashboard.html", context)
import openpyxl
from django.http import HttpResponse
from urllib.parse import urlencode

def export_excel(request):
    # On réutilise la logique de filtre de suivi_caisse
    response = suivi_caisse(request)
    ctx = response.context_data if hasattr(response, "context_data") else response.context

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mouvements caisse"

    headers = ["Date", "N° pièce", "Libellé", "Type", "Caisse",
               "Mode", "Entrée", "Sortie", "Solde cumulé"]
    ws.append(headers)

    for ligne in ctx["lignes"]:
        m = ligne["mouvement"]
        ws.append([
            m.date_mouvement.strftime("%Y-%m-%d"),
            m.num_piece,
            m.libelle,
            m.type_mouvement.libelle,
            m.caisse.libelle,
            m.mode_paiement.libelle,
            m.entree or 0,
            m.sortie or 0,
            ligne["solde_cumule"],
        ])

    # Réponse HTTP
    response_http = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response_http["Content-Disposition"] = 'attachment; filename="suivi_caisse.xlsx"'
    wb.save(response_http)
    return response_http

def export_pdf(request):
    """
    Génère un PDF du suivi de caisse en tenant compte des mêmes filtres
    que la page /caisse/suivi/.
    """
    # On réutilise la logique de filtres de suivi_caisse
    date_debut_str = request.GET.get("date_debut")
    date_fin_str = request.GET.get("date_fin")
    caisse_id = request.GET.get("caisse")
    type_id = request.GET.get("type")
    mode_id = request.GET.get("mode")

    date_debut = parse_date(date_debut_str) if date_debut_str else None
    date_fin = parse_date(date_fin_str) if date_fin_str else None

    param = Parametre.objects.order_by("-annee_analyse").first()
    solde_initial = param.solde_initial if param else 0
    solde_courant = solde_initial

    mouvements = MouvementCaisse.objects.all()

    if date_debut:
        mouvements = mouvements.filter(date_mouvement__gte=date_debut)
    if date_fin:
        mouvements = mouvements.filter(date_mouvement__lte=date_fin)
    if caisse_id:
        mouvements = mouvements.filter(caisse_id=caisse_id)
    if type_id:
        mouvements = mouvements.filter(type_mouvement_id=type_id)
    if mode_id:
        mouvements = mouvements.filter(mode_paiement_id=mode_id)

    mouvements = mouvements.select_related(
        "caisse", "type_mouvement", "mode_paiement"
    ).order_by("date_mouvement", "id")

    lignes = []
    total_entree = 0
    total_sortie = 0

    for mvt in mouvements:
        entree = mvt.entree or 0
        sortie = mvt.sortie or 0
        solde_courant += entree - sortie
        total_entree += entree
        total_sortie += sortie
        lignes.append(
            {
                "mouvement": mvt,
                "solde_cumule": solde_courant,
            }
        )

    context = {
        "parametre": param,
        "solde_initial": solde_initial,
        "lignes": lignes,
        "date_debut": date_debut_str or "",
        "date_fin": date_fin_str or "",
        "total_entree": total_entree,
        "total_sortie": total_sortie,
        "solde_final": solde_courant,
    }

    template = get_template("caisse/pdf_suivi_caisse.html")
    html = template.render(context)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="suivi_caisse.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Erreur lors de la génération du PDF", status=500)

    return response

def _filtrer_mouvements_queryset(request):
    """
    Réutilise la logique de filtres (même que dans suivi_caisse)
    et renvoie (queryset, params_filtres).
    """
    date_debut = request.GET.get("date_debut") or ""
    date_fin = request.GET.get("date_fin") or ""
    caisse_id = request.GET.get("caisse") or ""
    type_id = request.GET.get("type") or ""
    mode_id = request.GET.get("mode") or ""
    q = request.GET.get("q") or ""

    qs = (
        MouvementCaisse.objects
        .select_related("caisse", "type_mouvement", "mode_paiement")
        .order_by("date_mouvement", "id")  # pour solde cumulatif
    )

    if date_debut:
        qs = qs.filter(date_mouvement__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_mouvement__lte=date_fin)
    if caisse_id:
        qs = qs.filter(caisse_id=caisse_id)
    if type_id:
        qs = qs.filter(type_mouvement_id=type_id)
    if mode_id:
        qs = qs.filter(mode_paiement_id=mode_id)
    if q:
        qs = qs.filter(
            Q(libelle__icontains=q)
            | Q(num_piece__icontains=q)
            | Q(observations__icontains=q)
        )

    params = {
        "date_debut": date_debut,
        "date_fin": date_fin,
        "caisse_id": caisse_id,
        "type_id": type_id,
        "mode_id": mode_id,
        "q": q,
    }
    return qs, params


def export_pdf(request):
    """
    Génère une page HTML 'relevé de caisse' prête à être imprimée en PDF
    (via Ctrl+P -> Imprimer en PDF).
    On pourra plus tard brancher un vrai moteur PDF si tu veux.
    """
    mouvements, params = _filtrer_mouvements_queryset(request)

    # Paramètres (solde initial)
    param = Parametre.objects.first()
    solde_initial = param.solde_initial if param else 0

    # Calcul du solde cumulé
    solde = solde_initial
    lignes = []
    total_entree = 0
    total_sortie = 0

    for m in mouvements:
        if m.entree:
            solde += m.entree
            total_entree += m.entree
        if m.sortie:
            solde -= m.sortie
            total_sortie += m.sortie
        lignes.append({
            "mouvement": m,
            "solde_cumule": solde,
        })

    solde_final = solde

    context = {
        "lignes": lignes,
        "total_entree": total_entree,
        "total_sortie": total_sortie,
        "solde_initial": solde_initial,
        "solde_final": solde_final,
        "date_debut": params["date_debut"],
        "date_fin": params["date_fin"],
        "now": timezone.now(),
        "request": request,  # pour {{ request.user.username }}
    }

    # On réutilise ton ancien template 'style PDF'
    return render(request, "caisse/pdf_suivi_caisse.html", context)


def export_excel(request):
    """
    Export Excel (XLSX) du relevé de caisse, avec les mêmes filtres que l'écran.
    """
    mouvements, params = _filtrer_mouvements_queryset(request)

    # Paramètres (solde initial)
    param = Parametre.objects.first()
    solde_initial = param.solde_initial if param else 0

    # Préparation du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relevé de caisse"

    # Titre + infos
    ws["A1"] = "Relevé de caisse"
    ws["A2"] = f"Période : {params['date_debut'] or 'toutes'} → {params['date_fin'] or 'toutes'}"
    ws["A3"] = f"Solde initial : {solde_initial} FCFA"
    ws["A4"] = f"Généré le : {timezone.now().strftime('%d/%m/%Y %H:%M')}"

    # En-têtes du tableau (ligne 6)
    headers = [
        "Date",
        "N° pièce",
        "Libellé",
        "Type",
        "Caisse",
        "Mode",
        "Entrée",
        "Sortie",
        "Solde cumulé",
    ]
    start_row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Contenu
    solde = solde_initial
    row = start_row + 1
    total_entree = 0
    total_sortie = 0

    for m in mouvements:
        if m.entree:
            solde += m.entree
            total_entree += m.entree
        if m.sortie:
            solde -= m.sortie
            total_sortie += m.sortie

        ws.cell(row=row, column=1, value=m.date_mouvement.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=2, value=m.num_piece or "")
        ws.cell(row=row, column=3, value=m.libelle)
        ws.cell(row=row, column=4, value=m.type_mouvement.libelle)
        ws.cell(row=row, column=5, value=m.caisse.libelle)
        ws.cell(row=row, column=6, value=m.mode_paiement.libelle)
        ws.cell(row=row, column=7, value=float(m.entree or 0))
        ws.cell(row=row, column=8, value=float(m.sortie or 0))
        ws.cell(row=row, column=9, value=float(solde))
        row += 1

    # Totaux en bas
    ws.cell(row=row + 1, column=6, value="Total entrées")
    ws.cell(row=row + 1, column=7, value=float(total_entree))
    ws.cell(row=row + 2, column=6, value="Total sorties")
    ws.cell(row=row + 2, column=8, value=float(total_sortie))
    ws.cell(row=row + 3, column=6, value="Solde final")
    ws.cell(row=row + 3, column=9, value=float(solde))

    # Largeur des colonnes
    col_widths = [12, 12, 40, 15, 15, 15, 12, 12, 15]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = "releve_caisse.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
